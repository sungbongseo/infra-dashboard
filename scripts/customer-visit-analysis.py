"""
거래처 방문 정밀 분석 — 채권 협상 자료.

대상 (정확한 명칭):
  - 대성이앤씨 주식회사 (코드 C100002940)
  - 건진케미컬 (코드 C100020396)

사용자 지칭 "대성이엔씨/건진케미칼" → 실제 표기 차이 확인됨.
"""
import openpyxl
from collections import defaultdict
from datetime import datetime

DATA_DIR = "기타/업로드자료"
TARGETS = {
    "대성이앤씨": ["대성이앤씨", "대성이엔씨"],  # 검색 패턴
    "건진케미컬": ["건진케미컬", "건진케미칼"],
}

MONTH_LABELS = ["2026-01", "2025-12", "2025-11", "2025-10", "2025-09", "2025-08", "2025-08 이전"]
LEDGER_COLS = [7, 10, 13, 16, 19, 22, 25]
SHIPMENT_COLS = [6, 9, 12, 15, 18, 21, 24]
TXN_COLS = [8, 11, 14, 17, 20, 23, 26]

def matches_target(cust_str: str, patterns) -> bool:
    cust_lower = cust_str.lower()
    return any(p.lower() in cust_lower for p in patterns)

print("=" * 80)
print("📋 거래처 방문 정밀 분석 — 채권 협상용")
print("기준일: 2026-04-27")
print(f"대상: {' / '.join(TARGETS.keys())}")
print("=" * 80)

# ─────────────────────────────────────────────────────────────
# 1. 미수채권 aging
# ─────────────────────────────────────────────────────────────
aging_files = [
    "건자재_미수채권연령.xlsx",
    "광주사무소_미수채권연령.xlsx",
    "대구사무소_미수채권연령.xlsx",
    "대전사무소_미수채권연령.xlsx",
    "부산지점_미수채권연령.xlsx",
    "전략구매혁신팀_미수채권연령.xlsx",
    "해외사업팀_미수채권연령.xlsx",
]

# {target: [{office, person, code, ledger by month, ...}]}
aging_data = defaultdict(list)

for f in aging_files:
    path = f"{DATA_DIR}/{f}"
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 5:
                continue
            cust_name = row[4]
            if cust_name is None:
                continue
            cust_str = str(cust_name).strip()

            for tgt_key, patterns in TARGETS.items():
                if matches_target(cust_str, patterns):
                    record = {
                        "name": cust_str,
                        "office": f.replace("_미수채권연령.xlsx", ""),
                        "org": str(row[1]) if row[1] else "",
                        "person": str(row[2]) if row[2] else "",
                        "code": str(row[3]) if row[3] else "",
                        "currency": str(row[5]) if row[5] else "",
                        "ledger": {},
                        "shipment": {},
                        "txn": {},
                    }
                    for i, lbl in enumerate(MONTH_LABELS):
                        record["ledger"][lbl] = float(row[LEDGER_COLS[i]] or 0) if len(row) > LEDGER_COLS[i] else 0
                        record["shipment"][lbl] = float(row[SHIPMENT_COLS[i]] or 0) if len(row) > SHIPMENT_COLS[i] else 0
                        record["txn"][lbl] = float(row[TXN_COLS[i]] or 0) if len(row) > TXN_COLS[i] else 0
                    record["total_shipment"] = float(row[27] or 0) if len(row) > 27 else 0
                    record["total_txn"] = float(row[28] or 0) if len(row) > 28 else 0
                    record["total_ledger"] = float(row[29] or 0) if len(row) > 29 else 0
                    record["credit_limit"] = float(row[30] or 0) if len(row) > 30 else 0

                    aging_data[tgt_key].append(record)
                    break
        wb.close()
    except Exception as e:
        print(f"⚠ {f}: {e}")

# ─────────────────────────────────────────────────────────────
# 2. 100 손익 데이터 (매출/품목/마진)
# ─────────────────────────────────────────────────────────────
print("\n[데이터 수집] 100 손익 분석 중 (48,369 행)...")

wb = openpyxl.load_workbook(f"{DATA_DIR}/100거래처별,품목별 손익.xlsx", read_only=True, data_only=True)
ws = wb.active

last_org = ""
last_cust = ""
last_item = ""
last_category = ""
last_month = ""

# {target: {item: {monthly: {YYYYMM: {sales, qty, profit, cost}}, totals}}}
profit_data = defaultdict(lambda: defaultdict(lambda: {
    "monthly": defaultdict(lambda: {"sales": 0, "qty": 0, "profit": 0, "cost": 0}),
    "total_sales": 0,
    "total_profit": 0,
    "total_qty": 0,
    "total_cost": 0,
    "category": "",
    "name": "",
}))
profit_monthly = defaultdict(lambda: defaultdict(lambda: {"sales": 0, "qty": 0, "profit": 0, "cost": 0}))

for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or row[0] is None:
        continue

    org = row[3] if row[3] else last_org
    cust_code = row[4] if row[4] else last_cust
    item = row[5] if row[5] else last_item
    category = row[6] if row[6] else last_category
    month_raw = row[13] if row[13] else last_month

    last_org, last_cust, last_item, last_category, last_month = org, cust_code, item, category, month_raw

    if not cust_code:
        continue
    cust_str = str(cust_code).strip()

    matched_target = None
    for tgt_key, patterns in TARGETS.items():
        if matches_target(cust_str, patterns):
            matched_target = tgt_key
            break
    if not matched_target:
        continue

    qty = float(row[43] or 0)
    sales = float(row[46] or 0)
    cost = float(row[49] or 0) if len(row) > 49 else 0
    profit = float(row[76] or 0) if len(row) > 76 else 0

    month_str = str(month_raw).strip()[:6]
    if not month_str.isdigit() or len(month_str) != 6:
        continue
    if any(s in str(item) for s in ["합계", "소계", "총계"]):
        continue

    item_str = str(item)
    profit_data[matched_target][item_str]["name"] = item_str
    profit_data[matched_target][item_str]["category"] = category or profit_data[matched_target][item_str]["category"]
    profit_data[matched_target][item_str]["monthly"][month_str]["sales"] += sales
    profit_data[matched_target][item_str]["monthly"][month_str]["qty"] += qty
    profit_data[matched_target][item_str]["monthly"][month_str]["profit"] += profit
    profit_data[matched_target][item_str]["monthly"][month_str]["cost"] += cost
    profit_data[matched_target][item_str]["total_sales"] += sales
    profit_data[matched_target][item_str]["total_profit"] += profit
    profit_data[matched_target][item_str]["total_qty"] += qty
    profit_data[matched_target][item_str]["total_cost"] += cost

    profit_monthly[matched_target][month_str]["sales"] += sales
    profit_monthly[matched_target][month_str]["qty"] += qty
    profit_monthly[matched_target][month_str]["profit"] += profit
    profit_monthly[matched_target][month_str]["cost"] += cost

wb.close()

# ─────────────────────────────────────────────────────────────
# 3. 보고서 출력
# ─────────────────────────────────────────────────────────────
def fmt_won(v):
    """원 → 억/천만/만원 표기"""
    if abs(v) >= 1e8: return f"{v/1e8:,.2f}억"
    if abs(v) >= 1e4: return f"{v/1e4:,.0f}만"
    return f"{v:,.0f}원"

for tgt_key in TARGETS:
    print(f"\n\n{'━'*80}")
    print(f"🏢 {tgt_key.upper()}")
    print(f"{'━'*80}")

    aging_records = aging_data.get(tgt_key, [])

    # 3-1. 미수채권
    print(f"\n📋 [미수채권 aging]")
    if not aging_records:
        print("   ❌ 발견 안 됨")
    else:
        # 통합 합계
        total_ledger = sum(r["total_ledger"] for r in aging_records)
        total_credit = sum(r["credit_limit"] for r in aging_records)
        total_shipment = sum(r["total_shipment"] for r in aging_records)
        total_txn = sum(r["total_txn"] for r in aging_records)

        print(f"   거래처 발견: {len(aging_records)}건 (다른 사무소/담당 분산)")
        for r in aging_records:
            ledger_sum = r['total_ledger']
            limit = r['credit_limit']
            usage = (ledger_sum / limit * 100) if limit else 0
            usage_marker = ""
            if usage >= 100: usage_marker = " 🚨 한도 초과"
            elif usage >= 90: usage_marker = " 🚨 한도 임박 (90%+)"
            elif usage >= 80: usage_marker = " ⚠ 한도 80%+"

            print(f"\n   ▶ [{r['code']}] {r['name']}")
            print(f"     사무소: {r['office']} · 담당: {r['person']}")
            print(f"     미수(장부): {fmt_won(ledger_sum):>12} / 한도 {fmt_won(limit):>10} = 사용률 {usage:.1f}%{usage_marker}")

            # Aging 분포 (장부 = 미수)
            print(f"     월별 미수 분포:")
            for lbl in MONTH_LABELS:
                ledger = r['ledger'][lbl]
                ship = r['shipment'][lbl]
                if ship == 0 and ledger == 0:
                    continue
                aging_marker = ""
                if "이전" in lbl: aging_marker = " 🚨 장기연체 (8M+)"
                elif lbl == "2025-08": aging_marker = " ⚠ 8M 이상"
                elif lbl == "2025-09": aging_marker = " ⚠ 7M"
                print(f"       {lbl:<13} 출고:{fmt_won(ship):>10}  미수:{fmt_won(ledger):>10}{aging_marker}")

        # 통합 메트릭
        usage_total = (total_ledger / total_credit * 100) if total_credit else 0
        print(f"\n   📊 통합 합계:")
        print(f"     총 미수 (장부)    : {fmt_won(total_ledger)}")
        print(f"     총 여신한도       : {fmt_won(total_credit)}")
        print(f"     총 사용률         : {usage_total:.1f}%")
        print(f"     누적 출고 (전기간): {fmt_won(total_shipment)}")
        print(f"     누적 거래 (전기간): {fmt_won(total_txn)}")

        # 장기연체 위험 점수
        long_overdue = sum(r['ledger']["2025-08 이전"] for r in aging_records)
        if long_overdue > 0:
            pct = long_overdue / total_ledger * 100 if total_ledger else 0
            print(f"\n   🚨 장기연체 위험: {fmt_won(long_overdue)} ({pct:.1f}% of 총 미수)")

    # 3-2. 100 손익 분석
    items_data = profit_data.get(tgt_key, {})
    monthly_data = profit_monthly.get(tgt_key, {})

    print(f"\n📈 [매출/품목/마진 추이 — 100 손익]")
    if not items_data:
        print("   ❌ 100 데이터에서 거래 발견 안 됨")
    else:
        sorted_months = sorted(monthly_data.keys())

        # 월별 추이
        print(f"\n   월별 추이 ({len(sorted_months)}개월):")
        print(f"   {'월':<10} {'매출':>15} {'영업이익':>15} {'마진율':>8} {'수량':>10}")
        total_s = total_p = total_q = 0
        for m in sorted_months:
            d = monthly_data[m]
            margin = (d['profit'] / d['sales'] * 100) if d['sales'] else 0
            margin_marker = ""
            if margin < 0: margin_marker = " 🚨 적자"
            elif margin < 5: margin_marker = " ⚠ 저마진"
            elif margin > 30: margin_marker = " 💰 고수익"
            print(f"   {m:<10} {fmt_won(d['sales']):>15} {fmt_won(d['profit']):>15} {margin:>7.1f}% {d['qty']:>10,.0f}{margin_marker}")
            total_s += d['sales']
            total_p += d['profit']
            total_q += d['qty']

        avg_margin = (total_p / total_s * 100) if total_s else 0
        print(f"   {'─'*65}")
        print(f"   {'합계':<10} {fmt_won(total_s):>15} {fmt_won(total_p):>15} {avg_margin:>7.1f}% {total_q:>10,.0f}")

        # Top 품목
        sorted_items = sorted(items_data.items(), key=lambda x: x[1]["total_sales"], reverse=True)[:10]
        print(f"\n   📦 Top 10 거래 품목:")
        print(f"   {'#':>2} {'품목':<35} {'카테고리':<10} {'매출':>12} {'마진율':>7} {'거래월':>5}")
        for i, (item_code, d) in enumerate(sorted_items, 1):
            margin = (d["total_profit"] / d["total_sales"] * 100) if d["total_sales"] else 0
            months_count = sum(1 for m in d["monthly"].values() if m["sales"] > 0)
            cat = (d["category"] or "")[:10]
            item_disp = item_code[:35]
            margin_marker = ""
            if margin < 0: margin_marker = " 🚨"
            elif margin > 30: margin_marker = " 💰"
            print(f"   {i:>2} {item_disp:<35} {cat:<10} {fmt_won(d['total_sales']):>12} {margin:>6.1f}%{margin_marker} {months_count:>5}")

        # 추이 분석
        if len(sorted_months) >= 6:
            recent_3m = sum(monthly_data[m]['sales'] for m in sorted_months[-3:])
            prev_3m = sum(monthly_data[m]['sales'] for m in sorted_months[-6:-3])
            recent_3m_p = sum(monthly_data[m]['profit'] for m in sorted_months[-3:])
            prev_3m_p = sum(monthly_data[m]['profit'] for m in sorted_months[-6:-3])
            growth = (recent_3m / prev_3m - 1) * 100 if prev_3m else 0
            growth_p = (recent_3m_p / prev_3m_p - 1) * 100 if prev_3m_p else 0

            print(f"\n   💹 추이 (최근 3M vs 직전 3M):")
            print(f"     매출:    {fmt_won(prev_3m):>10} → {fmt_won(recent_3m):>10} ({growth:+.1f}%)")
            print(f"     영업이익: {fmt_won(prev_3m_p):>10} → {fmt_won(recent_3m_p):>10} ({growth_p:+.1f}%)")

print("\n\n" + "=" * 80)
print("정밀 분석 완료")
print("=" * 80)
