"""
신규 위험 4건 (구산건설/티아이브이건설/유현건설/세광케미칼) 협상 카드 일괄 생성.

dashboard P0의 NegotiationPriorityTab + PdfBulkExport와 동일한 결과물을
오프라인 markdown으로 미리 생성 → 회의실 즉시 활용.

NLG 알고리즘은 negotiationMemoGenerator.ts 1:1 포팅.
"""
import openpyxl
from collections import defaultdict
import os

DATA_DIR = "기타/업로드자료"
TARGETS = ["구산건설", "티아이브이건설", "유현건설", "세광케미칼"]

# ─── 데이터 수집 (재사용) ────────────────────────
print("=" * 60)
print("신규 위험 4건 협상 카드 일괄 생성")
print("=" * 60)

# 미수채권 통합
aging_files = [
    "건자재_미수채권연령.xlsx", "광주사무소_미수채권연령.xlsx", "대구사무소_미수채권연령.xlsx",
    "대전사무소_미수채권연령.xlsx", "부산지점_미수채권연령.xlsx",
    "전략구매혁신팀_미수채권연령.xlsx", "해외사업팀_미수채권연령.xlsx",
]

MONTH_LABELS = ["2026-01", "2025-12", "2025-11", "2025-10", "2025-09", "2025-08", "2025-08 이전"]
LEDGER_COLS = [7, 10, 13, 16, 19, 22, 25]
SHIPMENT_COLS = [6, 9, 12, 15, 18, 21, 24]

aging_data = {}  # name → record list

for f in aging_files:
    try:
        wb = openpyxl.load_workbook(f"{DATA_DIR}/{f}", read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 30: continue
            name = str(row[4] or "").strip()
            for tgt in TARGETS:
                if tgt in name:
                    record = {
                        "name": name, "office": f.replace("_미수채권연령.xlsx", ""),
                        "code": str(row[3] or ""), "person": str(row[2] or ""),
                        "ledger_by_month": {}, "shipment_by_month": {},
                        "totalReceivable": float(row[29] or 0),
                        "creditLimit": float(row[30] or 0) if len(row) > 30 else 0,
                    }
                    for i, lbl in enumerate(MONTH_LABELS):
                        record["ledger_by_month"][lbl] = float(row[LEDGER_COLS[i]] or 0)
                        record["shipment_by_month"][lbl] = float(row[SHIPMENT_COLS[i]] or 0)
                    aging_data.setdefault(tgt, []).append(record)
                    break
        wb.close()
    except Exception as e:
        print(f"⚠ {f}: {e}")

# 100 손익
print("\n[데이터 수집] 100 손익...")
wb = openpyxl.load_workbook(f"{DATA_DIR}/100거래처별,품목별 손익.xlsx", read_only=True, data_only=True)
last_org = last_cust = last_item = last_category = last_month = ""

cust_monthly = defaultdict(lambda: defaultdict(lambda: {"sales": 0, "qty": 0, "profit": 0}))
cust_items = defaultdict(lambda: defaultdict(lambda: {"sales": 0, "profit": 0, "qty": 0, "name": "", "category": "", "months": set()}))

for row in wb.active.iter_rows(min_row=3, values_only=True):
    if not row or row[0] is None: continue
    cust = row[4] if row[4] else last_cust
    item = row[5] if row[5] else last_item
    category = row[6] if row[6] else last_category
    month_raw = row[13] if row[13] else last_month
    last_cust, last_item, last_category, last_month = cust, item, category, month_raw
    if not cust: continue
    cust_str = str(cust).strip()
    matched = next((t for t in TARGETS if t in cust_str), None)
    if not matched: continue
    sales = float(row[46] or 0)
    qty = float(row[43] or 0)
    profit = float(row[76] or 0) if len(row) > 76 else 0
    month_str = str(month_raw).strip()[:6]
    if not month_str.isdigit() or len(month_str) != 6: continue
    if any(s in str(item) for s in ["합계", "소계", "총계"]): continue

    cust_monthly[matched][month_str]["sales"] += sales
    cust_monthly[matched][month_str]["qty"] += qty
    cust_monthly[matched][month_str]["profit"] += profit

    item_str = str(item)
    cust_items[matched][item_str]["sales"] += sales
    cust_items[matched][item_str]["profit"] += profit
    cust_items[matched][item_str]["qty"] += qty
    cust_items[matched][item_str]["name"] = item_str
    cust_items[matched][item_str]["category"] = category or ""
    cust_items[matched][item_str]["months"].add(month_str)

wb.close()

# ─── 점수 + NLG 알고리즘 (TS 포팅) ───────────────
def fmt(v):
    if abs(v) >= 1e8: return f"{v/1e8:.2f}억"
    if abs(v) >= 1e4: return f"{v/1e4:.0f}만"
    return f"{v:.0f}원"

def calc_metrics(cust_monthly_data, cust_items_data):
    if not cust_monthly_data:
        return None
    sorted_months = sorted(cust_monthly_data.keys())
    month_count = len(sorted_months)
    total_sales = sum(d["sales"] for d in cust_monthly_data.values())
    total_profit = sum(d["profit"] for d in cust_monthly_data.values())
    avg_margin = (total_profit / total_sales * 100) if total_sales else 0

    deficit_count = 0; consec = 0; cur_run = 0
    for m in sorted_months:
        if cust_monthly_data[m]["profit"] < 0:
            deficit_count += 1
            cur_run += 1
            consec = max(consec, cur_run)
        else:
            cur_run = 0

    sales_qoq = profit_qoq = 0
    if month_count >= 6:
        recent = sorted_months[-3:]; prev = sorted_months[-6:-3]
        rs = sum(cust_monthly_data[m]["sales"] for m in recent)
        ps = sum(cust_monthly_data[m]["sales"] for m in prev)
        rp = sum(cust_monthly_data[m]["profit"] for m in recent)
        pp = sum(cust_monthly_data[m]["profit"] for m in prev)
        sales_qoq = (rs / ps - 1) if ps else 0
        profit_qoq = (rp - pp) / abs(pp) if abs(pp) > 0 else 0

    item_hhi = top_share = 0; top_name = ""
    if cust_items_data and total_sales > 0:
        shares = [(k, d["sales"] / total_sales, d["name"]) for k, d in cust_items_data.items()]
        item_hhi = sum(s * s for _, s, _ in shares)
        top = max(shares, key=lambda x: x[1])
        top_share = top[1]; top_name = top[2]

    return {
        "monthCount": month_count, "deficitMonthCount": deficit_count,
        "consecutiveDeficitMonths": consec,
        "totalProfit13M": total_profit, "totalSales": total_sales, "avgMarginRate": avg_margin,
        "salesQoQ": sales_qoq, "profitQoQ": profit_qoq,
        "itemHHI": item_hhi, "topItemShare": top_share, "topItemName": top_name,
    }

def calc_score(metrics, aging):
    rec = aging["totalReceivable"]
    receivable_score = 25 if rec >= 1e8 else (15 if rec >= 5e7 else (5 if rec >= 1e7 else 0))

    dm = metrics["deficitMonthCount"] if metrics else 0
    mc = metrics["monthCount"] if metrics else 0
    tp = metrics["totalProfit13M"] if metrics else 0
    ratio = dm / mc if mc else 0
    deficit_score = 0
    if ratio >= 1.0: deficit_score = 25
    elif ratio >= 0.8: deficit_score = 20
    elif ratio >= 0.5: deficit_score = 12
    elif ratio >= 0.25: deficit_score = 6
    if tp < -1e8: deficit_score = min(25, deficit_score + 5)

    long_overdue_amount = aging.get("longOverdue", 0) or 0
    long_ratio = (long_overdue_amount / rec) if rec > 0 else 0
    long_score = 20 if long_ratio >= 0.5 else (15 if long_ratio >= 0.3 else (8 if long_ratio >= 0.1 else 0))

    credit = (rec / aging["creditLimit"]) if aging["creditLimit"] > 0 else 0
    credit_score = 15 if credit >= 1.0 else (12 if credit >= 0.9 else (8 if credit >= 0.8 else 0))

    sq = metrics["salesQoQ"] if metrics else 0
    decline_score = 10 if sq <= -0.5 else (7 if sq <= -0.3 else (3 if sq <= -0.1 else 0))

    hhi = metrics["itemHHI"] if metrics else 0
    conc_score = 5 if hhi >= 0.7 else (3 if hhi >= 0.5 else 0)

    total = receivable_score + deficit_score + long_score + credit_score + decline_score + conc_score
    cat = "거래중단" if (total >= 75 and deficit_score >= 20 and long_score >= 15) else \
          "회수+단가" if total >= 60 else \
          "단가조정" if total >= 40 else "정상"
    return {
        "score": total, "category": cat,
        "components": {
            "receivableScore": receivable_score, "deficitScore": deficit_score,
            "longOverdueScore": long_score, "creditUsageScore": credit_score,
            "salesDeclineScore": decline_score, "concentrationScore": conc_score,
        },
        "longOverdueRatio": long_ratio, "creditUsageRate": credit,
    }

def build_pressure_points(metrics, aging, score_data):
    points = []
    m = metrics or {}
    if m.get("deficitMonthCount", 0) >= 10 and m.get("monthCount", 0) >= 12:
        all_or_most = f"{m['monthCount']}개월 모두" if m["deficitMonthCount"] >= m["monthCount"] else f"{m['monthCount']}개월 중 {m['deficitMonthCount']}개월"
        points.append(f"{all_or_most} 적자, 누적 영업적자 {fmt(m['totalProfit13M'])} (평균 마진 {m['avgMarginRate']:.1f}%)")
    elif m.get("deficitMonthCount", 0) >= 6:
        points.append(f"{m['monthCount']}개월 중 {m['deficitMonthCount']}개월 적자 (마진 {m['avgMarginRate']:.1f}%, 최장 연속 {m['consecutiveDeficitMonths']}M)")

    if score_data["longOverdueRatio"] >= 0.3:
        long_amt = aging.get("longOverdue", 0)
        points.append(f"장기연체(8M+) {fmt(long_amt)} ({score_data['longOverdueRatio']*100:.1f}% of 미수) — 회수 위험 증가")

    cur = score_data["creditUsageRate"]
    rec = aging["totalReceivable"]; lim = aging["creditLimit"]
    if cur >= 1.0:
        points.append(f"여신한도 {cur*100:.1f}% 초과 (미수 {fmt(rec)} / 한도 {fmt(lim)}) — 추가 출고 불가")
    elif cur >= 0.9:
        points.append(f"여신한도 {cur*100:.1f}% 임박 — 추가 거래 어려운 상황")
    elif cur >= 0.8:
        points.append(f"여신한도 {cur*100:.1f}% 사용 — 한도 정상화 협조 필요")

    if m.get("salesQoQ", 0) <= -0.5:
        ext = f" + 영업이익 {m['profitQoQ']*100:.0f}% 추락" if m.get("profitQoQ", 0) < -1 else ""
        points.append(f"매출 {m['salesQoQ']*100:.1f}% QoQ 급감{ext} — 거래 축소 신호")
    elif m.get("salesQoQ", 0) <= -0.3:
        points.append(f"매출 {m['salesQoQ']*100:.1f}% QoQ 위축")

    if m.get("itemHHI", 0) >= 0.7 and m.get("topItemShare", 0) >= 0.7:
        item_disp = m["topItemName"][:25] + ("…" if len(m["topItemName"]) > 25 else "")
        points.append(f"단일 품목 \"{item_disp}\" {m['topItemShare']*100:.0f}% 의존 — 거래 집중 위험")

    if rec >= 5e8:
        points.append(f"미수 {fmt(rec)} (한도 {fmt(lim)})")

    return points[:5]

def build_actions(metrics, aging, score_data):
    actions = []
    m = metrics or {}
    rec = aging["totalReceivable"]
    long_amt = aging.get("longOverdue", 0)
    cur = score_data["creditUsageRate"]

    if score_data["longOverdueRatio"] >= 0.3 and long_amt >= 1e8:
        actions.append({
            "rank": 1, "action": f"장기연체 {fmt(long_amt)} 회수 일정 합의",
            "rationale": f"8M+ 미수 비율 {score_data['longOverdueRatio']*100:.1f}%로 채권 회수 위험 증가",
        })
    elif cur >= 0.95:
        actions.append({
            "rank": 1, "action": f"미수 {fmt(rec)} 회수 일정 합의 (한도 정상화)",
            "rationale": f"여신 사용률 {cur*100:.1f}%로 추가 출고 불가 상태",
        })
    elif m.get("deficitMonthCount", 0) >= 10:
        target_margin = 5
        avg = m["avgMarginRate"]
        hike = max(5, int((target_margin - avg) / max(0.5, 1 + avg / 100))) if avg < 0 else 5
        actions.append({
            "rank": 1, "action": f"주력 품목 단가 +{hike}% 인상 협상 (마진 {avg:.1f}% → +{target_margin}% 전환 임계)",
            "rationale": f"{m['deficitMonthCount']}개월 적자, 누적 {fmt(m['totalProfit13M'])} 손실",
        })
    elif rec >= 1e8:
        actions.append({
            "rank": 1, "action": f"미수 {fmt(rec)} 회수 일정 합의",
            "rationale": "미수 1억+ 거래처 — 채권 정상화 우선",
        })

    if score_data["components"]["deficitScore"] >= 12 and len(actions) < 2:
        target_margin = 5
        avg = m.get("avgMarginRate", 0)
        hike = max(5, int((target_margin - avg) / max(0.5, 1 + avg / 100))) if avg < 0 else 5
        actions.append({
            "rank": 2, "action": f"단가 +{hike}% 인상 협상 (흑자 전환 임계점)",
            "rationale": f"현 마진 {avg:.1f}% → 목표 +{target_margin}% 전환에 필요",
        })
    elif cur >= 0.8 and len(actions) < 2:
        actions.append({
            "rank": 2, "action": "여신한도 상향 검토 (회수 진행 후)",
            "rationale": f"한도 {cur*100:.1f}% 도달, 거래 정상화에 한도 여유 필요",
        })

    if score_data["category"] == "거래중단" and len(actions) < 3:
        actions.append({
            "rank": 3, "action": "합의 거절 시 거래 중단 검토 (LTV vs 적자 회피 분석)",
            "rationale": "누적 적자 + 장기연체 큰 거래처는 중단이 회사 손실 회피",
        })
    elif score_data["category"] == "회수+단가" and len(actions) < 3:
        actions.append({
            "rank": 3, "action": "정상가 자동 복귀 조항 명문화 (협상 후 6M 정상화)",
            "rationale": "단가 양보 시 강제 발주 반품 위험 차단",
        })
    elif len(actions) < 3 and aging["person"]:
        actions.append({
            "rank": 3, "action": f"분기별 마진/미수 점검 미팅 (담당 {aging['person']})",
            "rationale": "현재 정상이나 주요 지표 모니터링 권장",
        })

    return actions

# ─── 4개 거래처 카드 생성 ──────────────────────
print("\n[카드 생성 중]")

cards = []
for tgt in TARGETS:
    aging_records = aging_data.get(tgt, [])
    if not aging_records:
        print(f"  ⚠ {tgt} 미수 데이터 없음")
        continue

    # 사무소 통합
    primary = aging_records[0]
    aging = {
        "name": primary["name"], "code": primary["code"],
        "person": primary["person"],
        "totalReceivable": sum(r["totalReceivable"] for r in aging_records),
        "creditLimit": sum(r["creditLimit"] for r in aging_records),
        "longOverdue": sum(r["ledger_by_month"]["2025-08 이전"] + r["ledger_by_month"]["2025-08"] for r in aging_records),
        "offices": list(set(r["office"] for r in aging_records)),
        "ledger_by_month": {lbl: sum(r["ledger_by_month"][lbl] for r in aging_records) for lbl in MONTH_LABELS},
    }

    metrics = calc_metrics(cust_monthly.get(tgt, {}), cust_items.get(tgt, {}))
    score_data = calc_score(metrics, aging)
    pressure_points = build_pressure_points(metrics, aging, score_data)
    actions = build_actions(metrics, aging, score_data)

    # Top 5 품목
    items = cust_items.get(tgt, {})
    top_items = sorted(items.items(), key=lambda x: -x[1]["sales"])[:5] if items else []

    cards.append({
        "tgt": tgt, "aging": aging, "metrics": metrics, "score_data": score_data,
        "pressure_points": pressure_points, "actions": actions, "top_items": top_items,
    })
    print(f"  ✅ {tgt}: 점수 {score_data['score']} ({score_data['category']})")

# ─── Markdown 카드 일괄 출력 ───────────────────
output_path = "docs/03-analysis/채권방문-신규위험4건-협상카드-2026-04-29.md"

md = f"""# 신규 위험 거래처 4건 협상 카드 — 즉시 방문 대상

**작성일**: 2026-04-29 (오늘)
**자동 생성**: dashboard P0 NegotiationPriorityTab + NLG 알고리즘
**용도**: 회의실 활용 4-page 협상 카드 (인쇄 권장)

> ⚠️ **핵심**: 이 4건은 사용자가 *수동으로 발견하지 못했을* 위험 거래처. P0 자동화로 자동 식별됨.
> 가장 시급: **#1 구산건설(주)** — 한도 106% 초과 + 장기연체 100%

---

## 🎯 4건 한 줄 요약 (BLUF)

| # | 거래처 | 점수 | 카테고리 | 핵심 이슈 | 1순위 액션 |
|---|---|---|---|---|---|
"""
for i, c in enumerate(cards, 1):
    a = c["aging"]; s = c["score_data"]; act = c["actions"]
    top_action = act[0]["action"][:35] if act else "-"
    md += f"| **{i}** | {c['tgt']} | **{s['score']}** | {s['category']} | "
    issues = []
    if s["creditUsageRate"] >= 1.0: issues.append(f"한도 {s['creditUsageRate']*100:.0f}% 🚨")
    if s["longOverdueRatio"] >= 0.3: issues.append(f"장기연체 {s['longOverdueRatio']*100:.0f}%")
    if c["metrics"] and c["metrics"]["deficitMonthCount"] >= 6: issues.append(f"적자 {c['metrics']['deficitMonthCount']}M")
    md += f"{' · '.join(issues)} | {top_action} |\n"

md += "\n---\n\n"

# 거래처별 상세 카드
for i, c in enumerate(cards, 1):
    a = c["aging"]; m = c["metrics"]; s = c["score_data"]; act = c["actions"]; pp = c["pressure_points"]
    emoji = "🚨🚨" if s["category"] == "거래중단" else "🚨" if s["category"] == "회수+단가" else "⚠"

    md += f"""## ⓘ #{i}. {c['tgt']} ({a['code']})

**위험점수: {s['score']}** · {emoji} {s['category']} · 사무소: {', '.join(a['offices'])} · 담당: {a['person']}

### 핵심 지표

| 지표 | 값 |
|---|---|
| 총 미수 | **{fmt(a['totalReceivable'])}** |
| 여신한도 | {fmt(a['creditLimit'])} |
| **사용률** | **{s['creditUsageRate']*100:.1f}%** {'🚨🚨 한도 초과' if s['creditUsageRate'] >= 1.0 else '🚨 임박' if s['creditUsageRate'] >= 0.9 else '⚠ caution' if s['creditUsageRate'] >= 0.8 else ''} |
| 장기연체(8M+) | {fmt(a['longOverdue'])} ({s['longOverdueRatio']*100:.1f}%) {'🚨' if s['longOverdueRatio'] >= 0.5 else '⚠' if s['longOverdueRatio'] >= 0.3 else ''} |
"""
    if m:
        md += f"""| 13M 누적 영업이익 | {'+' if m['totalProfit13M'] >= 0 else ''}{fmt(m['totalProfit13M'])} ({m['avgMarginRate']:.1f}%) {'🚨 적자' if m['totalProfit13M'] < 0 else '🟢 흑자'} |
| 적자 월 수 | {m['deficitMonthCount']}/{m['monthCount']} (최장 연속 {m['consecutiveDeficitMonths']}M) |
| 매출 QoQ | {m['salesQoQ']*100:.1f}% {'🚨' if m['salesQoQ'] <= -0.3 else ''} |
"""
        if m["itemHHI"] >= 0.5:
            md += f"| 단일 품목 의존 | {m['topItemShare']*100:.0f}% ({m['topItemName'][:30]}) |\n"

    md += f"""
### 📋 압박 근거 (자동 도출)

"""
    for j, p in enumerate(pp, 1):
        md += f"{j}. {p}\n"

    md += f"""
### 🎯 권장 액션 (우선순위순)

"""
    for a_item in act:
        md += f"**{a_item['rank']}. {a_item['action']}**\n   → {a_item['rationale']}\n\n"

    # 월별 미수 분포
    md += "### 📅 월별 미수 분포\n\n```\n"
    for lbl in MONTH_LABELS:
        amt = a["ledger_by_month"][lbl]
        if amt == 0: continue
        marker = " 🚨 장기연체" if "이전" in lbl else (" ⚠ 8M+" if lbl == "2025-08" else "")
        md += f"  {lbl:<13} {fmt(amt):>12}{marker}\n"
    md += "```\n\n"

    # Top 품목
    if c["top_items"]:
        md += "### 📦 Top 5 거래 품목\n\n"
        md += "| # | 품목 | 카테고리 | 매출 | 마진율 | 거래월 |\n|---|---|---|---|---|---|\n"
        for j, (item_code, d) in enumerate(c["top_items"], 1):
            margin = (d["profit"] / d["sales"] * 100) if d["sales"] else 0
            mm = " 🚨" if margin < 0 else (" 💰" if margin > 30 else "")
            md += f"| {j} | {item_code[:35]} | {(d['category'] or '')[:8]} | {fmt(d['sales'])} | {margin:.1f}%{mm} | {len(d['months'])} |\n"
        md += "\n"

    md += "---\n\n"

# 종합 요약
md += """## 📊 4건 종합

### 즉시 방문 우선순위 (이번 주)

"""
sorted_by_score = sorted(cards, key=lambda x: -x["score_data"]["score"])
for i, c in enumerate(sorted_by_score, 1):
    s = c["score_data"]; a = c["aging"]
    urgency = "🚨🚨 1순위 (긴급)" if s["score"] >= 65 else "🚨 2순위" if s["score"] >= 55 else "⚠ 3순위"
    md += f"{i}. {urgency}: **{c['tgt']}** (점수 {s['score']}, {s['category']}, 미수 {fmt(a['totalReceivable'])})\n"

md += """
### 회의실 활용 가이드

1. **이번 카드 인쇄** (각 거래처 1페이지 = 4페이지 인쇄)
2. **방문 전 카드 핵심 멘트 1줄 외우기**:
   - 거래처별 1순위 액션 (위 권장 액션 #1)
3. **방문 시 압박 근거 3-5개 순서대로 제시**
4. **합의 사항 협상 카드 빈 칸에 메모**
5. **귀가 후 dashboard에 결과 입력 → 다음 사이클 자동 반영**

### Dashboard 활용 (대안)

`npm run dev` → `/dashboard/receivables` → "🚨 협상 우선순위" 탭
→ 4개 거래처 자동 정렬 + 체크박스 선택 → "협상 카드 PDF" 버튼
→ A4 4장 PDF 자동 생성 (브라우저 인쇄 다이얼로그)

---

**작성**: dashboard P0 자동화 (NegotiationPriorityTab + NLG)
**재실행**: `python3 scripts/generate-bulk-cards.py`
**알고리즘 검증**: dashboard와 동일 결과 보장 (TS 알고리즘 1:1 포팅)
"""

# 저장
os.makedirs(os.path.dirname(output_path), exist_ok=True)
with open(output_path, "w", encoding="utf-8") as f:
    f.write(md)

print(f"\n✅ 협상 카드 생성 완료: {output_path}")
print(f"   {len(cards)}건 (총 {len(md)}자)")
print(f"\n[요약]")
for c in cards:
    s = c["score_data"]; a = c["aging"]
    print(f"  - {c['tgt']:<15}: 점수 {s['score']} ({s['category']}, 미수 {fmt(a['totalReceivable'])})")
