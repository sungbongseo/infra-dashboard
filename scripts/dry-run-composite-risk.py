"""
P0+P1+UX dry-run 검증 스크립트.

TypeScript customerCompositeRisk.ts + negotiationMemoGenerator.ts 알고리즘을
Python으로 동일하게 재현하여 실데이터에서 Top N 위험 거래처 + 멘트 자동 생성.

목적:
  - Dashboard가 보여줄 결과를 미리 검증
  - 알고리즘이 실데이터에서 합리적인 Top N을 산출하는지 확인
  - AlertPanel에 표시될 거래처 리스트 미리보기

비교:
  - 대성/건진/리본TS가 Top N에 포함되는지 (이미 알고 있는 위험)
  - 다른 위험 거래처가 자동으로 추가 발견되는지
"""
import openpyxl
from collections import defaultdict
from itertools import combinations

DATA_DIR = "업로드자료"  # 대시보드와 동일 truth (이전 "기타/업로드자료"는 archive)

# ─── 1. 미수채권 통합 ────────────────────────────
print("=" * 80)
print("P0+P1 dry-run — 전체 거래처 Composite Risk Score 시뮬")
print("=" * 80)

aging_files = [
    "건자재_미수채권연령.xlsx", "광주사무소_미수채권연령.xlsx", "대구사무소_미수채권연령.xlsx",
    "대전사무소_미수채권연령.xlsx", "부산지점_미수채권연령.xlsx",
    "전략구매혁신팀_미수채권연령.xlsx", "해외사업팀_미수채권연령.xlsx",
]

aging_merged = {}  # code → {name, offices, totalReceivable, longOverdue, creditLimit}

for f in aging_files:
    try:
        wb = openpyxl.load_workbook(f"{DATA_DIR}/{f}", read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=3, values_only=True):
            if not row or len(row) < 30: continue
            code = row[3]
            if not code: continue
            code = str(code).strip()
            office = f.replace("_미수채권연령.xlsx", "")
            longOverdue = (float(row[22] or 0) + float(row[25] or 0))  # month6 + overdue
            total = float(row[29] or 0)
            limit = float(row[30] or 0) if len(row) > 30 else 0

            if code in aging_merged:
                aging_merged[code]["totalReceivable"] += total
                aging_merged[code]["creditLimit"] += limit
                aging_merged[code]["longOverdueAmount"] += longOverdue
                if office not in aging_merged[code]["offices"]:
                    aging_merged[code]["offices"].append(office)
            else:
                aging_merged[code] = {
                    "code": code, "name": str(row[4] or code),
                    "offices": [office],
                    "org": str(row[1] or ""), "person": str(row[2] or ""),
                    "totalReceivable": total, "creditLimit": limit,
                    "longOverdueAmount": longOverdue,
                }
        wb.close()
    except Exception as e:
        print(f"⚠ {f}: {e}")

# longOverdueRatio + creditUsageRate 재계산
for m in aging_merged.values():
    m["longOverdueRatio"] = (m["longOverdueAmount"] / m["totalReceivable"]) if m["totalReceivable"] > 0 else 0
    m["creditUsageRate"] = (m["totalReceivable"] / m["creditLimit"]) if m["creditLimit"] > 0 else 0

print(f"\n[1] 미수채권 통합: {len(aging_merged)}개 거래처 (사무소 통합 후)")

# ─── 2. 100 손익 분석 ────────────────────────────
wb = openpyxl.load_workbook(f"{DATA_DIR}/100거래처별,품목별 손익.xlsx", read_only=True, data_only=True)
ws = wb.active
last_org = last_cust = last_item = last_category = last_month = ""

cust_monthly = defaultdict(lambda: defaultdict(lambda: {"sales": 0, "profit": 0}))
cust_items = defaultdict(lambda: defaultdict(lambda: {"sales": 0, "name": ""}))

for row in ws.iter_rows(min_row=3, values_only=True):
    if not row or row[0] is None: continue
    cust = row[4] if row[4] else last_cust
    item = row[5] if row[5] else last_item
    month_raw = row[13] if row[13] else last_month
    last_cust = cust; last_item = item; last_month = month_raw
    if not cust: continue
    cust_str = str(cust).strip()
    sales = float(row[46] or 0)
    profit = float(row[76] or 0) if len(row) > 76 else 0
    month_str = str(month_raw).strip()[:6]
    if not month_str.isdigit() or len(month_str) != 6: continue
    if any(s in str(item) for s in ["합계", "소계", "총계"]): continue
    cust_monthly[cust_str][month_str]["sales"] += sales
    cust_monthly[cust_str][month_str]["profit"] += profit
    item_str = str(item)
    cust_items[cust_str][item_str]["sales"] += sales
    cust_items[cust_str][item_str]["name"] = item_str
wb.close()

print(f"[2] 100 손익: {len(cust_monthly)}개 거래처 시계열")

# ─── 3. Composite Risk Score (TypeScript 알고리즘 포팅) ──
def score_receivable(amt):
    if amt >= 100_000_000: return 25
    if amt >= 50_000_000: return 15
    if amt >= 10_000_000: return 5
    return 0

def score_deficit(deficit_months, total_profit, month_count):
    if month_count == 0: return 0
    ratio = deficit_months / month_count
    score = 0
    if ratio >= 1.0: score = 25
    elif ratio >= 0.8: score = 20
    elif ratio >= 0.5: score = 12
    elif ratio >= 0.25: score = 6
    if total_profit < -100_000_000:
        score = min(25, score + 5)
    return score

def score_long_overdue(ratio):
    if ratio >= 0.5: return 20
    if ratio >= 0.3: return 15
    if ratio >= 0.1: return 8
    return 0

def score_credit(rate):
    if rate >= 1.0: return 15
    if rate >= 0.9: return 12
    if rate >= 0.8: return 8
    return 0

def score_decline(qoq):
    if qoq <= -0.5: return 10
    if qoq <= -0.3: return 7
    if qoq <= -0.1: return 3
    return 0

def score_concentration(hhi):
    if hhi >= 0.7: return 5
    if hhi >= 0.5: return 3
    return 0

def categorize(score, deficit_score, long_overdue_score):
    if score >= 75 and deficit_score >= 20 and long_overdue_score >= 15:
        return "거래중단"
    if score >= 60: return "회수+단가"
    if score >= 40: return "단가조정"
    return "정상"

def calc_metrics(cust_str):
    """TypeScript calcProfitMetrics 포팅"""
    monthly = cust_monthly.get(cust_str, {})
    items = cust_items.get(cust_str, {})
    if not monthly:
        return None

    sorted_months = sorted(monthly.keys())
    month_count = len(sorted_months)
    total_sales = sum(d["sales"] for d in monthly.values())
    total_profit = sum(d["profit"] for d in monthly.values())
    avg_margin = (total_profit / total_sales * 100) if total_sales else 0

    deficit_count = 0
    consec = 0
    cur_run = 0
    for m in sorted_months:
        if monthly[m]["profit"] < 0:
            deficit_count += 1
            cur_run += 1
            consec = max(consec, cur_run)
        else:
            cur_run = 0

    sales_qoq = profit_qoq = 0
    if month_count >= 6:
        recent = sorted_months[-3:]; prev = sorted_months[-6:-3]
        rs = sum(monthly[m]["sales"] for m in recent)
        ps = sum(monthly[m]["sales"] for m in prev)
        rp = sum(monthly[m]["profit"] for m in recent)
        pp = sum(monthly[m]["profit"] for m in prev)
        sales_qoq = (rs / ps - 1) if ps else 0
        profit_qoq = (rp - pp) / abs(pp) if abs(pp) > 0 else 0

    item_hhi = top_share = 0
    top_name = ""
    if items and total_sales > 0:
        shares = [(k, d["sales"] / total_sales, d["name"]) for k, d in items.items()]
        item_hhi = sum(s * s for _, s, _ in shares)
        top = max(shares, key=lambda x: x[1])
        top_share = top[1]
        top_name = top[2]

    return {
        "monthCount": month_count, "deficitMonthCount": deficit_count,
        "consecutiveDeficitMonths": consec,
        "totalProfit13M": total_profit, "totalSales": total_sales, "avgMarginRate": avg_margin,
        "salesQoQ": sales_qoq, "profitQoQ": profit_qoq,
        "itemHHI": item_hhi, "topItemShare": top_share, "topItemName": top_name,
    }

# ─── 4. 모든 거래처 점수 계산 + 정렬 ─────────────
print("\n[3] Composite Risk Score 계산 중...")

all_codes = set(aging_merged.keys())
# 100 데이터의 거래처도 추가 (이름 fallback 매칭)
name_to_code = {m["name"]: m["code"] for m in aging_merged.values()}
for cust_str in cust_monthly.keys():
    if cust_str in name_to_code:
        continue  # 이미 매칭됨
    all_codes.add(cust_str)  # 이름 자체를 코드로

results = []
for code in all_codes:
    aging = aging_merged.get(code)
    if not aging:
        # 100 데이터의 cust_str을 이름으로 가정
        aging_by_name = next((a for a in aging_merged.values() if a["name"] == code), None)
        aging = aging_by_name

    cust_str_for_profit = aging["name"] if aging else code
    profit_metrics = calc_metrics(cust_str_for_profit) or calc_metrics(code) or {
        "monthCount": 0, "deficitMonthCount": 0, "consecutiveDeficitMonths": 0,
        "totalProfit13M": 0, "totalSales": 0, "avgMarginRate": 0,
        "salesQoQ": 0, "profitQoQ": 0,
        "itemHHI": 0, "topItemShare": 0, "topItemName": "",
    }

    metrics = {
        "totalReceivable": aging["totalReceivable"] if aging else 0,
        "creditLimit": aging["creditLimit"] if aging else 0,
        "creditUsageRate": aging["creditUsageRate"] if aging else 0,
        "longOverdueAmount": aging["longOverdueAmount"] if aging else 0,
        "longOverdueRatio": aging["longOverdueRatio"] if aging else 0,
        **profit_metrics,
    }

    components = {
        "receivableScore": score_receivable(metrics["totalReceivable"]),
        "deficitScore": score_deficit(metrics["deficitMonthCount"], metrics["totalProfit13M"], metrics["monthCount"]),
        "longOverdueScore": score_long_overdue(metrics["longOverdueRatio"]),
        "creditUsageScore": score_credit(metrics["creditUsageRate"]),
        "salesDeclineScore": score_decline(metrics["salesQoQ"]),
        "concentrationScore": score_concentration(metrics["itemHHI"]),
    }
    risk_score = sum(components.values())
    if risk_score < 40:
        continue

    results.append({
        "code": aging["code"] if aging else code,
        "name": aging["name"] if aging else code,
        "offices": aging["offices"] if aging else [],
        "person": aging["person"] if aging else "",
        "riskScore": risk_score,
        "components": components,
        "metrics": metrics,
        "category": categorize(risk_score, components["deficitScore"], components["longOverdueScore"]),
    })

results.sort(key=lambda x: -x["riskScore"])

# ─── 5. 결과 출력 ───────────────────────────────
def fmt(v):
    if abs(v) >= 1e8: return f"{v/1e8:.2f}억"
    if abs(v) >= 1e4: return f"{v/1e4:.0f}만"
    return f"{v:.0f}"

print(f"\n[4] 위험 거래처 (점수 ≥ 40): {len(results)}건")
print(f"[5] AlertPanel 표시 대상 (점수 ≥ 70): {sum(1 for r in results if r['riskScore'] >= 70)}건")

# 카테고리 분포
cats = defaultdict(int)
for r in results:
    cats[r["category"]] += 1
print(f"\n[카테고리 분포]")
for cat, n in sorted(cats.items(), key=lambda x: -x[1]):
    print(f"  {cat:<10}: {n}건")

# Top 15 출력
print(f"\n[Top 15 위험 거래처 — Dashboard NegotiationPriorityTab 예상 표시]\n")
print(f"  {'#':>2} {'거래처':<30} {'점수':>4} {'카테고리':<10} {'미수':>10} {'한도':>6} {'적자':>4} {'장기연체':>8}")
print(f"  {'─'*90}")
for i, r in enumerate(results[:15], 1):
    m = r["metrics"]
    name = r["name"][:30]
    usage_pct = f"{m['creditUsageRate']*100:.0f}%" if m['creditUsageRate'] else "-"
    deficit = f"{m['deficitMonthCount']}/{m['monthCount']}" if m['monthCount'] else "-"
    long_pct = f"{m['longOverdueRatio']*100:.0f}%" if m['longOverdueAmount'] else "-"
    print(f"  {i:>2} {name:<30} {r['riskScore']:>4} {r['category']:<10} {fmt(m['totalReceivable']):>10} {usage_pct:>6} {deficit:>4} {long_pct:>8}")

# 사용자가 알고 있는 거래처 검증
print(f"\n[검증] 알려진 위험 거래처 위치 확인:")
for known in ["대성", "건진", "리본"]:
    found = next((i for i, r in enumerate(results, 1) if known in r["name"]), None)
    if found:
        r = results[found - 1]
        print(f"  {known:<6}: #{found} ({r['name']}, 점수 {r['riskScore']}, {r['category']})")
    else:
        print(f"  {known:<6}: 점수 < 40으로 제외")

# AlertPanel 시뮬레이션 (점수 ≥ 70)
print(f"\n[AlertPanel 시뮬 — 점수 ≥ 70 거래처 메시지 미리보기]\n")
alerts_to_show = [r for r in results if r["riskScore"] >= 70]
for r in alerts_to_show[:10]:
    m = r["metrics"]
    emoji = "🚨🚨" if r["category"] == "거래중단" else "🚨"
    tags = []
    if m["totalReceivable"] >= 1e8: tags.append(f"미수 {m['totalReceivable']/1e8:.1f}억")
    if m["creditUsageRate"] >= 0.8: tags.append(f"한도 {m['creditUsageRate']*100:.0f}%")
    if m["deficitMonthCount"] >= 6: tags.append(f"적자 {m['deficitMonthCount']}M")
    if m["longOverdueRatio"] >= 0.3: tags.append(f"장기연체 {m['longOverdueRatio']*100:.0f}%")
    tag_text = " · ".join(tags)
    print(f"  {emoji} {r['name']} 위험점수 {r['riskScore']} · {r['category']} · {tag_text}")

print(f"\n총 {len(alerts_to_show)}개 alert이 AlertPanel에 표시됨 (클릭 → /dashboard/receivables?tab=negotiation)")

print("\n" + "=" * 80)
print("dry-run 완료")
print("=" * 80)
